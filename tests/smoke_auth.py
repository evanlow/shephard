"""Smoke tests for app/routes/auth.py

Covers: login page, login success/failure, dashboard auth guard,
logout, superuser-only user management, create/delete user flows.
"""

import sys
import os
import io
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app
from app.extensions import db
from app.models.user import User


def _make_app():
    app = create_app("testing")
    return app


def _create_user(username, email, password, is_superuser=False):
    user = User(username=username, email=email, is_superuser=is_superuser)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    return user


class TestLoginPage(unittest.TestCase):
    def setUp(self):
        self.app = _make_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def test_login_page_loads(self):
        resp = self.client.get("/login")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"Login", resp.data)

    def test_login_redirects_when_already_authenticated(self):
        _create_user("su", "su@test.com", "password123", is_superuser=True)
        self.client.post("/login", data={"username": "su", "password": "password123"})
        resp = self.client.get("/login")
        self.assertEqual(resp.status_code, 302)


class TestLoginPost(unittest.TestCase):
    def setUp(self):
        self.app = _make_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        _create_user("admin", "admin@test.com", "password123", is_superuser=True)
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def test_login_success_redirects(self):
        resp = self.client.post(
            "/login", data={"username": "admin", "password": "password123"}
        )
        self.assertEqual(resp.status_code, 302)

    def test_login_wrong_password_returns_401(self):
        resp = self.client.post(
            "/login", data={"username": "admin", "password": "wrongpassword"}
        )
        self.assertEqual(resp.status_code, 401)

    def test_login_unknown_user_returns_401(self):
        resp = self.client.post(
            "/login", data={"username": "nobody", "password": "password123"}
        )
        self.assertEqual(resp.status_code, 401)

    def test_login_missing_fields_returns_400(self):
        resp = self.client.post("/login", data={})
        self.assertEqual(resp.status_code, 400)


class TestDashboard(unittest.TestCase):
    def setUp(self):
        self.app = _make_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        _create_user("admin", "admin@test.com", "password123", is_superuser=True)
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def test_dashboard_requires_auth(self):
        resp = self.client.get("/dashboard")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login", resp.headers["Location"])
        self.assertIn("next=/dashboard", resp.headers["Location"])

    def test_dashboard_requires_auth_preserves_relative_query_next(self):
        resp = self.client.get("/dashboard?tab=weekly")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("next=/dashboard?tab%3Dweekly", resp.headers["Location"])

    def test_invalid_session_user_id_treated_as_anonymous(self):
        with self.client.session_transaction() as session:
            session["_user_id"] = "not-an-int"
        resp = self.client.get("/dashboard")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login", resp.headers["Location"])

    def test_dashboard_accessible_when_logged_in(self):
        self.client.post("/login", data={"username": "admin", "password": "password123"})
        resp = self.client.get("/dashboard")
        self.assertEqual(resp.status_code, 200)

    def test_logout_redirects(self):
        self.client.post("/login", data={"username": "admin", "password": "password123"})
        resp = self.client.post("/logout")
        self.assertEqual(resp.status_code, 302)


class TestUserManagement(unittest.TestCase):
    def setUp(self):
        self.app = _make_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        self.superuser = _create_user("su", "su@test.com", "password123", is_superuser=True)
        self.plain_user = _create_user("plain", "plain@test.com", "password123", is_superuser=False)
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def _login_as_superuser(self):
        self.client.post("/login", data={"username": "su", "password": "password123"})

    def _login_as_plain(self):
        self.client.post("/login", data={"username": "plain", "password": "password123"})

    def test_admin_users_requires_auth(self):
        resp = self.client.get("/admin/users")
        self.assertEqual(resp.status_code, 302)

    def test_admin_users_blocked_for_non_superuser(self):
        self._login_as_plain()
        resp = self.client.get("/admin/users")
        self.assertEqual(resp.status_code, 403)

    def test_admin_users_accessible_to_superuser(self):
        self._login_as_superuser()
        resp = self.client.get("/admin/users")
        self.assertEqual(resp.status_code, 200)

    def test_new_user_form_loads(self):
        self._login_as_superuser()
        resp = self.client.get("/admin/users/new")
        self.assertEqual(resp.status_code, 200)

    def test_create_user_success(self):
        self._login_as_superuser()
        resp = self.client.post(
            "/admin/users/new",
            data={
                "username": "newuser",
                "email": "new@test.com",
                "password": "validpass1",
                "confirm_password": "validpass1",
            },
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)
        created = db.session.execute(
            db.select(User).where(User.username == "newuser")
        ).scalar_one_or_none()
        self.assertIsNotNone(created)

    def test_create_user_short_password_returns_400(self):
        self._login_as_superuser()
        resp = self.client.post(
            "/admin/users/new",
            data={
                "username": "newuser",
                "email": "new@test.com",
                "password": "short",
                "confirm_password": "short",
            },
        )
        self.assertEqual(resp.status_code, 400)

    def test_create_user_password_mismatch_returns_400(self):
        self._login_as_superuser()
        resp = self.client.post(
            "/admin/users/new",
            data={
                "username": "newuser",
                "email": "new@test.com",
                "password": "password123",
                "confirm_password": "different123",
            },
        )
        self.assertEqual(resp.status_code, 400)

    def test_create_user_duplicate_username_returns_400(self):
        self._login_as_superuser()
        resp = self.client.post(
            "/admin/users/new",
            data={
                "username": "plain",
                "email": "unique@test.com",
                "password": "password123",
                "confirm_password": "password123",
            },
        )
        self.assertEqual(resp.status_code, 400)

    def test_delete_user_success(self):
        self._login_as_superuser()
        target_id = self.plain_user.id
        resp = self.client.post(
            f"/admin/users/{target_id}/delete", follow_redirects=False
        )
        self.assertEqual(resp.status_code, 302)
        deleted = db.session.get(User, target_id)
        self.assertIsNone(deleted)

    def test_cannot_delete_self(self):
        self._login_as_superuser()
        su_id = self.superuser.id
        resp = self.client.post(
            f"/admin/users/{su_id}/delete", follow_redirects=False
        )
        # Redirects back to list (not deleted)
        self.assertEqual(resp.status_code, 302)
        still_there = db.session.get(User, su_id)
        self.assertIsNotNone(still_there)


class TestSetUserPassword(unittest.TestCase):
    def setUp(self):
        self.app = _make_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        self.superuser = _create_user("su", "su@test.com", "password123", is_superuser=True)
        self.target = _create_user("target", "target@test.com", "oldpassword", is_superuser=False)
        self.plain_admin = _create_user("plain", "plain@test.com", "password123", is_superuser=False)
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def _login_as_superuser(self):
        self.client.post("/login", data={"username": "su", "password": "password123"})

    def _login_as_plain(self):
        self.client.post("/login", data={"username": "plain", "password": "password123"})

    def _get_password_form_csrf_token(self):
        resp = self.client.get(f"/admin/users/{self.target.id}/password")
        self.assertEqual(resp.status_code, 200)
        with self.client.session_transaction() as session:
            return session["set_user_password_csrf_token"]

    def test_password_form_loads_for_superuser(self):
        self._login_as_superuser()
        resp = self.client.get(f"/admin/users/{self.target.id}/password")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"Set Password", resp.data)
        self.assertIn(b'name="csrf_token"', resp.data)

    def test_password_form_blocked_for_non_superuser(self):
        self._login_as_plain()
        resp = self.client.get(f"/admin/users/{self.target.id}/password")
        self.assertEqual(resp.status_code, 403)

    def test_password_form_blocked_for_unauthenticated(self):
        resp = self.client.get(f"/admin/users/{self.target.id}/password")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login", resp.headers["Location"])

    def test_set_password_success(self):
        self._login_as_superuser()
        csrf_token = self._get_password_form_csrf_token()
        resp = self.client.post(
            f"/admin/users/{self.target.id}/password",
            data={
                "csrf_token": csrf_token,
                "password": "newpassword1",
                "confirm_password": "newpassword1",
            },
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)
        db.session.refresh(self.target)
        self.assertTrue(self.target.check_password("newpassword1"))

    def test_old_password_no_longer_works_after_reset(self):
        self._login_as_superuser()
        csrf_token = self._get_password_form_csrf_token()
        self.client.post(
            f"/admin/users/{self.target.id}/password",
            data={
                "csrf_token": csrf_token,
                "password": "newpassword1",
                "confirm_password": "newpassword1",
            },
        )
        db.session.refresh(self.target)
        self.assertFalse(self.target.check_password("oldpassword"))

    def test_new_password_works_after_reset(self):
        self._login_as_superuser()
        csrf_token = self._get_password_form_csrf_token()
        self.client.post(
            f"/admin/users/{self.target.id}/password",
            data={
                "csrf_token": csrf_token,
                "password": "newpassword1",
                "confirm_password": "newpassword1",
            },
        )
        # Log out and log back in with new password
        self.client.post("/logout")
        resp = self.client.post(
            "/login",
            data={"username": "target", "password": "newpassword1"},
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)

    def test_password_mismatch_returns_400(self):
        self._login_as_superuser()
        csrf_token = self._get_password_form_csrf_token()
        resp = self.client.post(
            f"/admin/users/{self.target.id}/password",
            data={
                "csrf_token": csrf_token,
                "password": "newpassword1",
                "confirm_password": "different123",
            },
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn(b"Passwords do not match.", resp.data)

    def test_short_password_returns_400(self):
        self._login_as_superuser()
        csrf_token = self._get_password_form_csrf_token()
        resp = self.client.post(
            f"/admin/users/{self.target.id}/password",
            data={
                "csrf_token": csrf_token,
                "password": "short",
                "confirm_password": "short",
            },
        )
        self.assertEqual(resp.status_code, 400)

    def test_missing_csrf_token_rejected(self):
        self._login_as_superuser()
        resp = self.client.post(
            f"/admin/users/{self.target.id}/password",
            data={"password": "newpassword1", "confirm_password": "newpassword1"},
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)
        db.session.refresh(self.target)
        self.assertTrue(self.target.check_password("oldpassword"))

    def test_post_blocked_for_non_superuser(self):
        self._login_as_plain()
        resp = self.client.post(
            f"/admin/users/{self.target.id}/password",
            data={"password": "newpassword1", "confirm_password": "newpassword1"},
        )
        self.assertEqual(resp.status_code, 403)

    def test_post_blocked_for_unauthenticated(self):
        resp = self.client.post(
            f"/admin/users/{self.target.id}/password",
            data={"password": "newpassword1", "confirm_password": "newpassword1"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login", resp.headers["Location"])

    def test_cannot_set_own_password_via_this_route(self):
        self._login_as_superuser()
        su_id = self.superuser.id
        resp = self.client.get(
            f"/admin/users/{su_id}/password", follow_redirects=True
        )
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"You cannot set your own password from this screen.", resp.data)

    def test_cannot_post_own_password_via_this_route(self):
        self._login_as_superuser()
        su_id = self.superuser.id
        resp = self.client.post(
            f"/admin/users/{su_id}/password",
            data={"password": "newpassword1", "confirm_password": "newpassword1"},
            follow_redirects=False,
        )
        self.assertEqual(resp.status_code, 302)


# ---------------------------------------------------------------------------
# System Purge
# ---------------------------------------------------------------------------

class TestSystemPurge(unittest.TestCase):
    def setUp(self):
        self.app = _make_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        self.superuser = _create_user("su", "su@test.com", "password123", is_superuser=True)
        self.plain_user = _create_user("plain", "plain@test.com", "password123", is_superuser=False)
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def _login_as_superuser(self):
        self.client.post("/login", data={"username": "su", "password": "password123"})

    def _login_as_plain(self):
        self.client.post("/login", data={"username": "plain", "password": "password123"})

    def _seed(self):
        """Create minimal data: one group, one member, one event, one attendance record."""
        from app.models.group import Group
        from app.models.member import Member
        from app.models.event import Event
        from app.models.attendance import Attendance
        from app.services.group_service import GroupService
        from datetime import datetime
        default = GroupService.get_default_group()
        group = Group(name="Worship")
        db.session.add(group)
        db.session.commit()
        member = Member(name="Test Person")
        db.session.add(member)
        db.session.commit()
        event = Event(name="Sunday", date=datetime(2026, 5, 1, 10, 0), group_id=group.id)
        db.session.add(event)
        db.session.commit()
        att = Attendance(event_id=event.id, member_id=member.id, present=True)
        db.session.add(att)
        db.session.commit()
        return group, member, event, att

    # --- access control ---

    def test_purge_page_requires_superuser(self):
        self._login_as_plain()
        resp = self.client.get("/admin/purge")
        self.assertEqual(resp.status_code, 403)

    def test_purge_page_requires_auth(self):
        resp = self.client.get("/admin/purge")
        self.assertEqual(resp.status_code, 302)

    def test_purge_page_loads_for_superuser(self):
        self._login_as_superuser()
        resp = self.client.get("/admin/purge")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"System Purge", resp.data)

    # --- confirmation validation ---

    def test_purge_attendance_wrong_confirm_redirects_with_error(self):
        self._login_as_superuser()
        self._seed()
        resp = self.client.post("/admin/purge/attendance", data={"confirm": "yes"})
        self.assertEqual(resp.status_code, 302)
        from app.models.attendance import Attendance
        count = db.session.query(Attendance).count()
        self.assertEqual(count, 1)  # not deleted

    # --- purge attendance ---

    def test_purge_attendance_deletes_all_records(self):
        self._login_as_superuser()
        self._seed()
        resp = self.client.post("/admin/purge/attendance", data={"confirm": "PURGE"})
        self.assertEqual(resp.status_code, 302)
        from app.models.attendance import Attendance
        self.assertEqual(db.session.query(Attendance).count(), 0)

    def test_purge_attendance_preserves_members_and_events(self):
        self._login_as_superuser()
        self._seed()
        self.client.post("/admin/purge/attendance", data={"confirm": "PURGE"})
        from app.models.member import Member
        from app.models.event import Event
        self.assertGreater(db.session.query(Member).count(), 0)
        self.assertGreater(db.session.query(Event).count(), 0)

    # --- purge members ---

    def test_purge_members_deletes_all_members(self):
        self._login_as_superuser()
        self._seed()
        resp = self.client.post("/admin/purge/members", data={"confirm": "PURGE"})
        self.assertEqual(resp.status_code, 302)
        from app.models.member import Member
        self.assertEqual(db.session.query(Member).count(), 0)

    def test_purge_members_also_deletes_attendance(self):
        self._login_as_superuser()
        self._seed()
        self.client.post("/admin/purge/members", data={"confirm": "PURGE"})
        from app.models.attendance import Attendance
        self.assertEqual(db.session.query(Attendance).count(), 0)

    def test_purge_members_preserves_groups_and_events(self):
        self._login_as_superuser()
        self._seed()
        self.client.post("/admin/purge/members", data={"confirm": "PURGE"})
        from app.models.group import Group
        from app.models.event import Event
        self.assertGreater(db.session.query(Group).count(), 0)
        self.assertGreater(db.session.query(Event).count(), 0)

    # --- purge events ---

    def test_purge_events_deletes_all_events(self):
        self._login_as_superuser()
        self._seed()
        resp = self.client.post("/admin/purge/events", data={"confirm": "PURGE"})
        self.assertEqual(resp.status_code, 302)
        from app.models.event import Event
        self.assertEqual(db.session.query(Event).count(), 0)

    def test_purge_events_also_deletes_attendance(self):
        self._login_as_superuser()
        self._seed()
        self.client.post("/admin/purge/events", data={"confirm": "PURGE"})
        from app.models.attendance import Attendance
        self.assertEqual(db.session.query(Attendance).count(), 0)

    def test_purge_events_preserves_members_and_groups(self):
        self._login_as_superuser()
        self._seed()
        self.client.post("/admin/purge/events", data={"confirm": "PURGE"})
        from app.models.member import Member
        from app.models.group import Group
        self.assertGreater(db.session.query(Member).count(), 0)
        self.assertGreater(db.session.query(Group).count(), 0)

    def test_purge_events_wrong_confirm_redirects_with_error(self):
        self._login_as_superuser()
        self._seed()
        resp = self.client.post("/admin/purge/events", data={"confirm": "yes"})
        self.assertEqual(resp.status_code, 302)
        from app.models.event import Event
        self.assertGreater(db.session.query(Event).count(), 0)

    # --- purge groups (original) ---

    def test_purge_groups_deletes_custom_groups(self):
        self._login_as_superuser()
        self._seed()
        resp = self.client.post("/admin/purge/groups", data={"confirm": "PURGE"})
        self.assertEqual(resp.status_code, 302)
        from app.models.group import Group
        from app.models.membership import DEFAULT_GROUP_NAME
        custom = db.session.query(Group).filter(Group.name != DEFAULT_GROUP_NAME).count()
        self.assertEqual(custom, 0)

    def test_purge_groups_preserves_all_members_group(self):
        self._login_as_superuser()
        self._seed()
        self.client.post("/admin/purge/groups", data={"confirm": "PURGE"})
        from app.models.group import Group
        from app.models.membership import DEFAULT_GROUP_NAME
        default = db.session.query(Group).filter(Group.name == DEFAULT_GROUP_NAME).first()
        self.assertIsNotNone(default)

    def test_purge_groups_also_deletes_events_and_attendance(self):
        self._login_as_superuser()
        self._seed()
        self.client.post("/admin/purge/groups", data={"confirm": "PURGE"})
        from app.models.event import Event
        from app.models.attendance import Attendance
        self.assertEqual(db.session.query(Event).count(), 0)
        self.assertEqual(db.session.query(Attendance).count(), 0)


# ---------------------------------------------------------------------------
# Full-system Excel export
# ---------------------------------------------------------------------------

class TestExport(unittest.TestCase):
    def setUp(self):
        self.app = _make_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        self.superuser = _create_user("su", "su@test.com", "password123", is_superuser=True)
        self.plain_user = _create_user("plain", "plain@test.com", "password123", is_superuser=False)
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def _login_as_superuser(self):
        self.client.post("/login", data={"username": "su", "password": "password123"})

    def _login_as_plain(self):
        self.client.post("/login", data={"username": "plain", "password": "password123"})

    def _seed(self):
        from app.models.group import Group
        from app.models.member import Member
        from app.models.event import Event
        from app.models.attendance import Attendance
        from app.services.group_service import GroupService
        from datetime import datetime
        GroupService.get_default_group()
        group = Group(name="Choir")
        db.session.add(group)
        db.session.commit()
        member = Member(name="Alice")
        db.session.add(member)
        db.session.commit()
        event = Event(name="Sunday Service", date=datetime(2026, 5, 4, 10, 0), group_id=group.id)
        db.session.add(event)
        db.session.commit()
        att = Attendance(event_id=event.id, member_id=member.id, present=True)
        db.session.add(att)
        db.session.commit()
        return group, member, event, att

    def test_export_requires_auth(self):
        resp = self.client.get("/admin/export")
        self.assertEqual(resp.status_code, 302)

    def test_export_requires_superuser(self):
        self._login_as_plain()
        resp = self.client.get("/admin/export")
        self.assertEqual(resp.status_code, 403)

    def test_export_returns_xlsx_for_superuser(self):
        self._login_as_superuser()
        resp = self.client.get("/admin/export")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp.content_type,
        )

    def test_export_is_attachment_with_filename(self):
        self._login_as_superuser()
        resp = self.client.get("/admin/export")
        disposition = resp.headers.get("Content-Disposition", "")
        self.assertIn("attachment", disposition)
        self.assertIn("shepherd_export_", disposition)
        self.assertIn(".xlsx", disposition)

    def test_export_first_sheet_is_members(self):
        import io
        import openpyxl
        self._login_as_superuser()
        self._seed()
        resp = self.client.get("/admin/export")
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        self.assertEqual(wb.sheetnames[0], "Members")

    def test_export_members_sheet_contains_member_data(self):
        import io
        import openpyxl
        self._login_as_superuser()
        self._seed()
        resp = self.client.get("/admin/export")
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        ws = wb["Members"]
        names = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        self.assertIn("Alice", names)

    def test_export_has_event_sheet(self):
        import io
        import openpyxl
        self._login_as_superuser()
        self._seed()
        resp = self.client.get("/admin/export")
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        # There should be more than just the Members sheet
        self.assertGreater(len(wb.sheetnames), 1)

    def test_export_empty_db_returns_only_members_sheet(self):
        import io
        import openpyxl
        self._login_as_superuser()
        resp = self.client.get("/admin/export")
        wb = openpyxl.load_workbook(io.BytesIO(resp.data))
        self.assertEqual(wb.sheetnames, ["Members"])


# ---------------------------------------------------------------------------
# Restore from backup
# ---------------------------------------------------------------------------

class TestRestore(unittest.TestCase):
    def setUp(self):
        self.app = _make_app()
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        self.superuser = _create_user("su", "su@test.com", "password123", is_superuser=True)
        self.plain_user = _create_user("plain", "plain@test.com", "password123", is_superuser=False)
        self.client = self.app.test_client()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def _login_as_superuser(self):
        self.client.post("/login", data={"username": "su", "password": "password123"})

    def _login_as_plain(self):
        self.client.post("/login", data={"username": "plain", "password": "password123"})

    def _seed_and_export(self):
        """Seed data, export it, then return the raw xlsx bytes."""
        from app.models.group import Group
        from app.models.member import Member
        from app.models.event import Event
        from app.models.attendance import Attendance
        from app.services.group_service import GroupService
        from datetime import datetime
        GroupService.get_default_group()
        group = Group(name="Choir")
        db.session.add(group)
        db.session.commit()
        member = Member(name="Alice")
        db.session.add(member)
        db.session.commit()
        # Add Alice to the Choir group so she is expected at Choir events
        member.groups.append(group)
        db.session.commit()
        event = Event(name="Sunday Service", date=datetime(2030, 1, 1, 10, 0), group_id=group.id)
        db.session.add(event)
        db.session.commit()
        att = Attendance(event_id=event.id, member_id=member.id, present=True)
        db.session.add(att)
        db.session.commit()
        resp = self.client.get("/admin/export")
        return resp.data

    # --- access control ---

    def test_restore_page_requires_auth(self):
        resp = self.client.get("/admin/restore")
        self.assertEqual(resp.status_code, 302)

    def test_restore_page_requires_superuser(self):
        self._login_as_plain()
        resp = self.client.get("/admin/restore")
        self.assertEqual(resp.status_code, 403)

    def test_restore_page_loads_for_superuser(self):
        self._login_as_superuser()
        resp = self.client.get("/admin/restore")
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"Restore", resp.data)

    def test_restore_upload_requires_superuser(self):
        self._login_as_plain()
        resp = self.client.post("/admin/restore", data={})
        self.assertEqual(resp.status_code, 403)

    # --- input validation ---

    def test_restore_no_file_redirects_with_error(self):
        self._login_as_superuser()
        resp = self.client.post("/admin/restore", data={}, follow_redirects=True)
        self.assertIn(b"No file selected", resp.data)

    def test_restore_wrong_extension_rejected(self):
        self._login_as_superuser()
        data = {"backup": (io.BytesIO(b"not an xlsx"), "export.csv")}
        resp = self.client.post("/admin/restore",
                                data=data, content_type="multipart/form-data",
                                follow_redirects=True)
        self.assertIn(b"xlsx", resp.data)

    def test_restore_corrupt_file_rejected(self):
        self._login_as_superuser()
        data = {"backup": (io.BytesIO(b"this is not a valid xlsx"), "export.xlsx")}
        resp = self.client.post("/admin/restore",
                                data=data, content_type="multipart/form-data",
                                follow_redirects=True)
        self.assertIn(b"Could not open file", resp.data)

    def test_restore_wrong_format_rejected(self):
        """A valid xlsx that doesn't match the export structure is rejected."""
        import io as _io
        import openpyxl
        self._login_as_superuser()
        wb = openpyxl.Workbook()
        wb.active.title = "WrongSheet"
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = {"backup": (buf, "export.xlsx")}
        resp = self.client.post("/admin/restore",
                                data=data, content_type="multipart/form-data",
                                follow_redirects=True)
        self.assertIn(b"Invalid file", resp.data)

    def test_restore_wrong_members_headers_rejected(self):
        """Export-shaped xlsx with wrong Members headers is rejected."""
        import io as _io
        import openpyxl
        self._login_as_superuser()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        ws.append(["#", "FullName", "Group"])   # wrong headers
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = {"backup": (buf, "export.xlsx")}
        resp = self.client.post("/admin/restore",
                                data=data, content_type="multipart/form-data",
                                follow_redirects=True)
        self.assertIn(b"Invalid file", resp.data)

    # --- full round-trip ---

    def test_restore_imports_members(self):
        """Export → purge members → restore → members are back."""
        from app.models.member import Member
        self._login_as_superuser()
        xlsx_bytes = self._seed_and_export()
        # Purge members
        self.client.post("/admin/purge/members", data={"confirm": "PURGE"})
        self.assertEqual(db.session.query(Member).count(), 0)
        # Restore
        data = {"backup": (io.BytesIO(xlsx_bytes), "export.xlsx")}
        resp = self.client.post("/admin/restore",
                                data=data, content_type="multipart/form-data",
                                follow_redirects=True)
        self.assertIn(b"Restore complete", resp.data)
        self.assertGreater(db.session.query(Member).count(), 0)

    def test_restore_imports_groups(self):
        """Custom groups are recreated after purge."""
        from app.models.group import Group
        from app.models.membership import DEFAULT_GROUP_NAME
        self._login_as_superuser()
        xlsx_bytes = self._seed_and_export()
        self.client.post("/admin/purge/groups", data={"confirm": "PURGE"})
        custom_before = db.session.query(Group).filter(Group.name != DEFAULT_GROUP_NAME).count()
        self.assertEqual(custom_before, 0)
        data = {"backup": (io.BytesIO(xlsx_bytes), "export.xlsx")}
        self.client.post("/admin/restore",
                         data=data, content_type="multipart/form-data")
        custom_after = db.session.query(Group).filter(Group.name != DEFAULT_GROUP_NAME).count()
        self.assertGreater(custom_after, 0)

    def test_restore_imports_events(self):
        """Events are recreated after purge groups (which cascades events)."""
        from app.models.event import Event
        self._login_as_superuser()
        xlsx_bytes = self._seed_and_export()
        self.client.post("/admin/purge/groups", data={"confirm": "PURGE"})
        self.assertEqual(db.session.query(Event).count(), 0)
        data = {"backup": (io.BytesIO(xlsx_bytes), "export.xlsx")}
        self.client.post("/admin/restore",
                         data=data, content_type="multipart/form-data")
        self.assertGreater(db.session.query(Event).count(), 0)

    def test_restore_imports_attendance(self):
        """Attendance records are restored after a full purge."""
        from app.models.attendance import Attendance
        self._login_as_superuser()
        xlsx_bytes = self._seed_and_export()
        # Full purge sequence: attendance → members → groups
        self.client.post("/admin/purge/attendance", data={"confirm": "PURGE"})
        self.client.post("/admin/purge/members", data={"confirm": "PURGE"})
        self.client.post("/admin/purge/groups", data={"confirm": "PURGE"})
        self.assertEqual(db.session.query(Attendance).count(), 0)
        data = {"backup": (io.BytesIO(xlsx_bytes), "export.xlsx")}
        self.client.post("/admin/restore",
                         data=data, content_type="multipart/form-data")
        self.assertGreater(db.session.query(Attendance).count(), 0)

    def test_restore_imports_attendance_when_row_index_is_float(self):
        """Google Sheets / Excel round-trips convert integer cells to floats;
        restore must still import attendance rows."""
        from openpyxl import load_workbook
        from app.models.attendance import Attendance
        self._login_as_superuser()
        xlsx_bytes = self._seed_and_export()

        # Simulate the Google Sheets round-trip: rewrite the row-index column
        # in every event sheet as a float (1 -> 1.0).
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        for sheet_name in wb.sheetnames[1:]:
            ws = wb[sheet_name]
            for r in range(7, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if isinstance(v, int) and not isinstance(v, bool):
                    ws.cell(row=r, column=1).value = float(v)
        buf = io.BytesIO()
        wb.save(buf)
        floatified = buf.getvalue()

        self.client.post("/admin/purge/attendance", data={"confirm": "PURGE"})
        self.client.post("/admin/purge/members", data={"confirm": "PURGE"})
        self.client.post("/admin/purge/groups", data={"confirm": "PURGE"})
        self.assertEqual(db.session.query(Attendance).count(), 0)

        data = {"backup": (io.BytesIO(floatified), "export.xlsx")}
        self.client.post("/admin/restore",
                         data=data, content_type="multipart/form-data")
        self.assertGreater(db.session.query(Attendance).count(), 0)

    def test_restore_member_is_in_all_members_group(self):
        """Restored members are enrolled in ALL MEMBERS."""
        from app.models.member import Member
        from app.models.membership import DEFAULT_GROUP_NAME
        self._login_as_superuser()
        xlsx_bytes = self._seed_and_export()
        self.client.post("/admin/purge/members", data={"confirm": "PURGE"})
        data = {"backup": (io.BytesIO(xlsx_bytes), "export.xlsx")}
        self.client.post("/admin/restore",
                         data=data, content_type="multipart/form-data")
        alice = db.session.execute(
            db.select(Member).where(Member.name == "Alice")
        ).scalar_one_or_none()
        self.assertIsNotNone(alice)
        group_names = [g.name for g in alice.groups]
        self.assertIn(DEFAULT_GROUP_NAME, group_names)

    # --- remarks / deactivation_reason export-restore round-trip ---

    def _seed_with_remarks_and_export(self):
        """Seed members with remarks and a deactivated member, then export."""
        from app.models.member import Member
        from app.services.group_service import GroupService
        from datetime import datetime
        GroupService.get_default_group()
        active = Member(name="Active Remarked", remarks="Husband of Mary")
        inactive = Member(
            name="Inactive Member",
            remarks="Long-time member",
            deactivated_at=datetime(2026, 5, 31, 23, 59, 59),
            deactivation_reason="Moved overseas",
        )
        db.session.add_all([active, inactive])
        db.session.commit()
        resp = self.client.get("/admin/export")
        self.assertEqual(resp.status_code, 200)
        return resp.data

    def test_export_includes_remarks_and_deactivation_reason_columns(self):
        import io as _io
        import openpyxl
        self._login_as_superuser()
        xlsx_bytes = self._seed_with_remarks_and_export()
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes))
        ws = wb["Members"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        self.assertEqual(headers[7], "Remarks")
        self.assertEqual(headers[8], "Deactivation Reason")
        # Inactive Member row should contain the values
        found = False
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == "Inactive Member":
                self.assertEqual(ws.cell(row=r, column=8).value, "Long-time member")
                self.assertEqual(ws.cell(row=r, column=9).value, "Moved overseas")
                found = True
        self.assertTrue(found)

    def test_restore_round_trips_remarks_and_deactivation_reason(self):
        from app.models.member import Member
        self._login_as_superuser()
        xlsx_bytes = self._seed_with_remarks_and_export()
        # Purge and restore
        self.client.post("/admin/purge/attendance", data={"confirm": "PURGE"})
        self.client.post("/admin/purge/members", data={"confirm": "PURGE"})
        data = {"backup": (io.BytesIO(xlsx_bytes), "export.xlsx")}
        self.client.post("/admin/restore",
                         data=data, content_type="multipart/form-data")
        restored = db.session.execute(
            db.select(Member).where(Member.name == "Inactive Member")
        ).scalar_one_or_none()
        self.assertIsNotNone(restored)
        self.assertEqual(restored.remarks, "Long-time member")
        self.assertEqual(restored.deactivation_reason, "Moved overseas")
        self.assertIsNotNone(restored.deactivated_at)
        active = db.session.execute(
            db.select(Member).where(Member.name == "Active Remarked")
        ).scalar_one_or_none()
        self.assertIsNotNone(active)
        self.assertEqual(active.remarks, "Husband of Mary")
        self.assertIsNone(active.deactivated_at)

    def test_restore_accepts_legacy_7_column_export(self):
        """Older backups without remarks/deactivation_reason columns are accepted."""
        import io as _io
        import openpyxl
        from app.models.member import Member
        self._login_as_superuser()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        ws.append(["#", "Name", "Primary Group", "All Groups", "Status", "Member Since", "Deactivated"])
        ws.append([1, "Legacy Member", "—", "ALL MEMBERS", "Active", "01 Jan 2024", ""])
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = {"backup": (buf, "export.xlsx")}
        resp = self.client.post("/admin/restore",
                                data=data, content_type="multipart/form-data",
                                follow_redirects=True)
        self.assertIn(b"Restore complete", resp.data)
        m = db.session.execute(
            db.select(Member).where(Member.name == "Legacy Member")
        ).scalar_one_or_none()
        self.assertIsNotNone(m)
        self.assertIsNone(m.remarks)
        self.assertIsNone(m.deactivation_reason)


if __name__ == "__main__":
    unittest.main()
