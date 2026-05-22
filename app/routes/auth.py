import io
from datetime import datetime, timezone
from functools import wraps

import openpyxl
from openpyxl.styles import Font

from flask import Blueprint, abort, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required, login_user, logout_user
from sqlalchemy import insert as sa_insert, text

from ..extensions import db
from ..models.attendance import Attendance
from ..models.event import Event
from ..models.group import Group
from ..models.member import Member
from ..models.membership import DEFAULT_GROUP_NAME, member_groups
from ..models.user import User
from ..services.attendance_service import AttendanceService

bp = Blueprint("auth", __name__)

# Expected structure constants used for both export validation and restore
_MEMBERS_HEADERS = ["#", "Name", "Primary Group", "All Groups", "Status", "Member Since", "Deactivated"]
_EVENT_META_LABELS = ["Event:", "Date:", "Group:", "Archived:"]


def admin_required(f):
    """Decorator: requires the current user to be an admin or superuser."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not (current_user.is_admin or current_user.is_superuser):
            abort(403)
        return f(*args, **kwargs)
    return decorated


def superuser_required(f):
    """Decorator: requires the current user to have is_superuser=True."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_superuser:
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# User management (superuser only)
# ---------------------------------------------------------------------------

@bp.get("/admin/users")
@superuser_required
def list_users():
    users = db.session.execute(db.select(User).order_by(User.username)).scalars().all()
    return render_template("auth/users.html", users=users)


@bp.get("/admin/users/new")
@superuser_required
def new_user():
    return render_template("auth/user_form.html")


@bp.post("/admin/users/new")
@superuser_required
def create_user():
    username = request.form.get("username", "").strip()
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "")
    confirm = request.form.get("confirm_password", "")

    errors = []
    if not username:
        errors.append("Username is required.")
    if not email:
        errors.append("Email is required.")
    if not password:
        errors.append("Password is required.")
    elif len(password) < 8:
        errors.append("Password must be at least 8 characters.")
    elif password != confirm:
        errors.append("Passwords do not match.")

    if not errors:
        existing = db.session.execute(
            db.select(User).where(
                (User.username == username) | (User.email == email)
            )
        ).scalar_one_or_none()
        if existing:
            errors.append("A user with that username or email already exists.")

    if errors:
        for e in errors:
            flash(e, "error")
        return render_template("auth/user_form.html",
                               username=username, email=email), 400

    user = User(username=username, email=email, is_admin=True)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    flash(f"User '{username}' created successfully.", "success")
    return redirect(url_for("auth.list_users"))


@bp.post("/admin/users/<int:user_id>/toggle-admin")
@superuser_required
def toggle_admin(user_id: int):
    if user_id == current_user.id:
        flash("You cannot change your own admin status.", "error")
        return redirect(url_for("auth.list_users"))

    user = db.session.get(User, user_id)
    if not user:
        flash("User not found.", "error")
        return redirect(url_for("auth.list_users"))

    if user.is_superuser:
        flash("Superuser accounts cannot be modified this way.", "error")
        return redirect(url_for("auth.list_users"))

    user.is_admin = not user.is_admin
    db.session.commit()

    status = "granted" if user.is_admin else "revoked"
    flash(f"Admin access {status} for '{user.username}'.", "success")
    return redirect(url_for("auth.list_users"))


@bp.post("/admin/users/<int:user_id>/delete")
@superuser_required
def delete_user(user_id: int):
    if user_id == current_user.id:
        flash("You cannot delete your own account.", "error")
        return redirect(url_for("auth.list_users"))

    user = db.session.get(User, user_id)
    if not user:
        flash("User not found.", "error")
        return redirect(url_for("auth.list_users"))

    db.session.delete(user)
    db.session.commit()
    flash(f"User '{user.username}' deleted.", "success")
    return redirect(url_for("auth.list_users"))


@bp.get("/login")
def login():
    if current_user.is_authenticated:
        return redirect(url_for("auth.dashboard"))
    return render_template("auth/login.html")


@bp.post("/login")
def login_post():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    remember = bool(request.form.get("remember"))

    if not username or not password:
        flash("Username and password are required.", "error")
        return render_template("auth/login.html"), 400

    user = db.session.execute(
        db.select(User).where(User.username == username)
    ).scalar_one_or_none()

    if user is None or not user.check_password(password):
        flash("Invalid username or password.", "error")
        return render_template("auth/login.html"), 401

    login_user(user, remember=remember)

    next_page = request.args.get("next")
    # Guard against open-redirect: only allow relative paths
    if next_page and next_page.startswith("/") and not next_page.startswith("//"):
        return redirect(next_page)
    return redirect(url_for("auth.dashboard"))


@bp.get("/dashboard")
@login_required
def dashboard():
    from ..services.event_service import EventService
    from ..services.group_service import GroupService
    from ..services.member_service import MemberService

    members = MemberService.get_all()
    groups = GroupService.get_all()
    events = EventService.get_all()
    recent_events = events[:5]

    return render_template(
        "auth/dashboard.html",
        member_count=len(members),
        group_count=len(groups),
        event_count=len(events),
        recent_events=recent_events,
    )


@bp.post("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("auth.login"))


# ---------------------------------------------------------------------------
# System Purge (superuser only)
# ---------------------------------------------------------------------------

@bp.get("/admin/purge")
@superuser_required
def purge_page():
    return render_template("auth/purge.html")


@bp.post("/admin/purge/attendance")
@superuser_required
def purge_attendance():
    if request.form.get("confirm", "").strip().upper() != "PURGE":
        flash("Type PURGE to confirm.", "error")
        return redirect(url_for("auth.purge_page"))
    count = db.session.query(Attendance).delete(synchronize_session=False)
    db.session.commit()
    flash(f"All attendance records deleted ({count} record{'s' if count != 1 else ''} removed).", "success")
    return redirect(url_for("auth.purge_page"))


@bp.post("/admin/purge/members")
@superuser_required
def purge_members():
    if request.form.get("confirm", "").strip().upper() != "PURGE":
        flash("Type PURGE to confirm.", "error")
        return redirect(url_for("auth.purge_page"))
    # Delete in FK order: attendance → member_groups junction → members
    db.session.query(Attendance).delete(synchronize_session=False)
    db.session.execute(text("DELETE FROM member_groups"))
    count = db.session.query(Member).delete(synchronize_session=False)
    db.session.commit()
    flash(f"All members deleted ({count} member{'s' if count != 1 else ''} removed).", "success")
    return redirect(url_for("auth.purge_page"))


@bp.post("/admin/purge/events")
@superuser_required
def purge_events():
    if request.form.get("confirm", "").strip().upper() != "PURGE":
        flash("Type PURGE to confirm.", "error")
        return redirect(url_for("auth.purge_page"))
    # Delete in FK order: attendance → events
    db.session.query(Attendance).delete(synchronize_session=False)
    count = db.session.query(Event).delete(synchronize_session=False)
    db.session.commit()
    flash(f"All events deleted ({count} event{'s' if count != 1 else ''} removed).", "success")
    return redirect(url_for("auth.purge_page"))


@bp.post("/admin/purge/groups")
@superuser_required
def purge_groups():
    if request.form.get("confirm", "").strip().upper() != "PURGE":
        flash("Type PURGE to confirm.", "error")
        return redirect(url_for("auth.purge_page"))
    # Collect non-default group IDs
    non_default_ids = db.session.execute(
        db.select(Group.id).where(Group.name != DEFAULT_GROUP_NAME)
    ).scalars().all()

    if non_default_ids:
        # Cascade manually: attendance for events → events → member_groups → groups
        event_ids = db.session.execute(
            db.select(Event.id).where(Event.group_id.in_(non_default_ids))
        ).scalars().all()
        if event_ids:
            db.session.query(Attendance).filter(
                Attendance.event_id.in_(event_ids)
            ).delete(synchronize_session=False)
        db.session.query(Event).filter(
            Event.group_id.in_(non_default_ids)
        ).delete(synchronize_session=False)
        db.session.execute(
            member_groups.delete().where(member_groups.c.group_id.in_(non_default_ids))
        )
        db.session.query(Member).filter(
            Member.group_id.in_(non_default_ids)
        ).update({"group_id": None}, synchronize_session=False)
        db.session.query(Group).filter(
            Group.id.in_(non_default_ids)
        ).delete(synchronize_session=False)
        db.session.commit()

    n = len(non_default_ids)
    flash(f"All custom groups deleted ({n} group{'s' if n != 1 else ''} removed; ALL MEMBERS preserved).", "success")
    return redirect(url_for("auth.purge_page"))


# ---------------------------------------------------------------------------
# Full-system Excel export (superuser only)
# ---------------------------------------------------------------------------

@bp.get("/admin/export")
@superuser_required
def export_all():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Members ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Members"

    header = ["#", "Name", "Primary Group", "All Groups", "Status", "Member Since", "Deactivated"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    members = db.session.execute(db.select(Member).order_by(Member.name)).scalars().all()
    for i, m in enumerate(members, 1):
        primary = m.group.name if m.group else "—"
        all_groups = ", ".join(g.name for g in m.groups)
        status = "Inactive" if m.deactivated_at else "Active"
        since = m.created_at.strftime("%d %b %Y") if m.created_at else "—"
        deactivated = m.deactivated_at.strftime("%d %b %Y") if m.deactivated_at else ""
        ws.append([i, m.name, primary, all_groups, status, since, deactivated])

    # Set column widths
    for col, width in zip("ABCDEFG", [5, 30, 20, 40, 12, 15, 15]):
        ws.column_dimensions[col].width = width

    # ── One sheet per event (oldest first) ───────────────────────────────────
    events = db.session.execute(
        db.select(Event).order_by(Event.date.asc(), Event.name)
    ).scalars().all()

    used_names = {"Members"}
    for event in events:
        # Build a unique sheet name within Excel's 31-char limit
        raw = f"{event.date.strftime('%d%b%y')} {event.name}"
        sheet_name = raw[:31]
        if sheet_name in used_names:
            base = raw[:28]
            n = 2
            while f"{base}({n})" in used_names:
                n += 1
            sheet_name = f"{base}({n})"
        used_names.add(sheet_name)

        ws_ev = wb.create_sheet(title=sheet_name)

        # Event metadata header block
        meta = [
            ("Event:", event.name),
            ("Date:", event.date.strftime("%d %b %Y")),
            ("Group:", event.group.name if event.group else "—"),
            ("Archived:", "Yes" if event.is_archived else "No"),
        ]
        for label, value in meta:
            ws_ev.append([label, value])
            ws_ev.cell(row=ws_ev.max_row, column=1).font = Font(bold=True)
        ws_ev.append([])  # blank separator

        # Attendance table header
        att_header_row = ws_ev.max_row + 1
        ws_ev.append(["#", "Name", "Present"])
        for cell in ws_ev[att_header_row]:
            cell.font = Font(bold=True)

        # Attendance data
        status_data, _ = AttendanceService.get_event_status(event.id)
        if status_data and status_data["expected_members"]:
            present_ids = {m["id"] for m in status_data["present_members"]}
            for j, m_data in enumerate(status_data["expected_members"], 1):
                present = "Yes" if m_data["id"] in present_ids else "No"
                ws_ev.append([j, m_data["name"], present])
            ws_ev.append([])
            ws_ev.append(["", "Present:", status_data["present_count"]])
            ws_ev.append(["", "Absent:", status_data["absent_count"]])
            ws_ev.append(["", "Total:", status_data["expected_count"]])
            for row in ws_ev.iter_rows(min_row=ws_ev.max_row - 2, max_row=ws_ev.max_row, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(bold=True)
        else:
            ws_ev.append(["", "(No eligible members for this event)", ""])

        ws_ev.column_dimensions["A"].width = 5
        ws_ev.column_dimensions["B"].width = 30
        ws_ev.column_dimensions["C"].width = 10

    # ── Stream to response ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"shepherd_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# Restore helpers
# ---------------------------------------------------------------------------

def _parse_restore_date(value):
    """Parse an exported date string (dd Mon YYYY) into a UTC datetime."""
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%d %b %Y").replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _validate_export_workbook(wb):
    """Return an error string if the workbook doesn't match the export format, else None."""
    if not wb.sheetnames or wb.sheetnames[0] != "Members":
        return "Invalid file: first sheet must be named 'Members'."
    ws = wb["Members"]
    actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
    if actual_headers != _MEMBERS_HEADERS:
        return "Invalid file: 'Members' sheet has unexpected column headers."
    for sheet_name in wb.sheetnames[1:]:
        ws_ev = wb[sheet_name]
        labels = [ws_ev.cell(row=r, column=1).value for r in range(1, 5)]
        if labels != _EVENT_META_LABELS:
            return f"Invalid file: sheet '{sheet_name}' does not match the expected event layout."
    return None


def _import_workbook(wb):
    """
    Import members, groups, events, and attendance from a validated export workbook.
    Uses Core-level INSERTs for members so the ORM after_insert listener is bypassed
    and we can set created_at / joined_at to the original values.
    Returns a summary dict.
    """
    from ..services.group_service import GroupService

    ws_members = wb["Members"]

    # ── Step 1: Collect every group name referenced anywhere in the file ──────
    all_group_names = set()
    for row in ws_members.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        if row[3]:
            for g in str(row[3]).split(","):
                g = g.strip()
                if g:
                    all_group_names.add(g)
    for sheet_name in wb.sheetnames[1:]:
        ws_ev = wb[sheet_name]
        group_cell = ws_ev.cell(row=3, column=2).value
        if group_cell:
            all_group_names.add(str(group_cell).strip())

    # ── Step 2: Ensure default group; create any missing custom groups ─────────
    default_group = GroupService.get_default_group()
    group_map = {DEFAULT_GROUP_NAME: default_group}

    for name in sorted(all_group_names):
        if name == DEFAULT_GROUP_NAME:
            continue
        existing = db.session.execute(
            db.select(Group).where(Group.name == name)
        ).scalar_one_or_none()
        if existing:
            group_map[name] = existing
        else:
            g = Group(name=name)
            db.session.add(g)
            db.session.flush()
            group_map[name] = g

    # ── Step 3: Import members via Core INSERT (bypasses ORM listener so we
    #            can set created_at and joined_at to their original values) ─────
    member_map = {}   # member name → member_id  (first occurrence wins)
    members_created = 0

    for row in ws_members.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        _, name, primary_group_name, all_groups_str, _status, member_since_str, deactivated_str = row
        if not name:
            continue

        name = str(name).strip()
        created_at = _parse_restore_date(member_since_str) or datetime.now(timezone.utc)

        deactivated_at = None
        if deactivated_str:
            d = _parse_restore_date(deactivated_str)
            if d:
                deactivated_at = d.replace(hour=23, minute=59, second=59, microsecond=0)

        pname = str(primary_group_name).strip() if primary_group_name else ""
        primary_group = group_map.get(pname) if pname and pname != "—" else None
        group_id = primary_group.id if primary_group and pname != DEFAULT_GROUP_NAME else None

        # Core INSERT → ORM after_insert event does NOT fire
        result = db.session.execute(
            sa_insert(Member.__table__).values(
                name=name,
                group_id=group_id,
                created_at=created_at,
                deactivated_at=deactivated_at,
            ).returning(Member.__table__.c.id)
        )
        member_id = result.scalar()

        # Build the member_groups rows with the original joined_at
        seen_gids = set()
        groups_to_add = []

        if all_groups_str:
            for g_name in str(all_groups_str).split(","):
                g_name = g_name.strip()
                if not g_name:
                    continue
                g = group_map.get(g_name)
                if g and g.id not in seen_gids:
                    groups_to_add.append({"member_id": member_id, "group_id": g.id, "joined_at": created_at})
                    seen_gids.add(g.id)

        # Always ensure ALL MEMBERS membership exists
        default_gid = default_group.id
        if default_gid not in seen_gids:
            groups_to_add.append({"member_id": member_id, "group_id": default_gid, "joined_at": created_at})

        db.session.execute(member_groups.insert(), groups_to_add)

        if name not in member_map:
            member_map[name] = member_id
        members_created += 1

    # ── Step 4: Import events and attendance ──────────────────────────────────
    events_created = 0
    attendance_created = 0

    for sheet_name in wb.sheetnames[1:]:
        ws_ev = wb[sheet_name]
        event_name = ws_ev.cell(row=1, column=2).value
        date_str   = ws_ev.cell(row=2, column=2).value
        group_name = ws_ev.cell(row=3, column=2).value
        archived_s = ws_ev.cell(row=4, column=2).value

        event_date = _parse_restore_date(date_str)
        if not event_date or not event_name or not group_name:
            continue

        group = group_map.get(str(group_name).strip())
        if not group:
            continue

        is_archived = str(archived_s).strip().lower() == "yes" if archived_s else False
        event = Event(name=str(event_name).strip(), date=event_date,
                      group_id=group.id, is_archived=is_archived)
        db.session.add(event)
        db.session.flush()
        events_created += 1

        # Attendance rows begin at row 7 (rows 1-4 meta, row 5 blank, row 6 header)
        for r in range(7, ws_ev.max_row + 1):
            num_val     = ws_ev.cell(row=r, column=1).value
            name_val    = ws_ev.cell(row=r, column=2).value
            present_val = ws_ev.cell(row=r, column=3).value
            # Accept int or float (Google Sheets / Excel round-trips convert ints to floats)
            is_row_index = isinstance(num_val, (int, float)) and not isinstance(num_val, bool)
            if not is_row_index:
                break   # blank row or summary section
            if not name_val:
                continue
            mid = member_map.get(str(name_val).strip())
            if mid is None:
                continue
            present = str(present_val).strip().lower() == "yes" if present_val else False
            db.session.add(Attendance(event_id=event.id, member_id=mid, present=present))
            attendance_created += 1

    db.session.commit()
    return {
        "groups": len([k for k in group_map if k != DEFAULT_GROUP_NAME]),
        "members": members_created,
        "events": events_created,
        "attendance": attendance_created,
    }


# ---------------------------------------------------------------------------
# Restore from backup (superuser only)
# ---------------------------------------------------------------------------

@bp.get("/admin/restore")
@superuser_required
def restore_page():
    has_data = (
        db.session.query(Member).count() > 0
        or db.session.query(Event).count() > 0
    )
    return render_template("auth/restore.html", has_data=has_data)


@bp.post("/admin/restore")
@superuser_required
def restore_upload():
    f = request.files.get("backup")
    if not f or not f.filename:
        flash("No file selected.", "error")
        return redirect(url_for("auth.restore_page"))

    if not f.filename.lower().endswith(".xlsx"):
        flash("Invalid file: only .xlsx spreadsheets are accepted.", "error")
        return redirect(url_for("auth.restore_page"))

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))
    except Exception:
        flash("Could not open file. Make sure it is a valid .xlsx spreadsheet.", "error")
        return redirect(url_for("auth.restore_page"))

    error = _validate_export_workbook(wb)
    if error:
        flash(error, "error")
        return redirect(url_for("auth.restore_page"))

    try:
        summary = _import_workbook(wb)
    except Exception as exc:
        db.session.rollback()
        flash(f"Restore failed: {exc}", "error")
        return redirect(url_for("auth.restore_page"))

    parts = [
        f"{summary['groups']} group{'s' if summary['groups'] != 1 else ''}",
        f"{summary['members']} member{'s' if summary['members'] != 1 else ''}",
        f"{summary['events']} event{'s' if summary['events'] != 1 else ''}",
        f"{summary['attendance']} attendance record{'s' if summary['attendance'] != 1 else ''}",
    ]
    flash(f"Restore complete: {', '.join(parts)} imported.", "success")
    return redirect(url_for("auth.restore_page"))
